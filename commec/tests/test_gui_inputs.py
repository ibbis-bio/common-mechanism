"""Input handling for the Commec GUI."""

import json
from io import BytesIO

import openpyxl
import pytest
import xlwt
import yaml

from commec.gui import server


@pytest.fixture
def client(tmp_path, monkeypatch):
    runs_dir = tmp_path / "runs"
    runs_dir.mkdir()
    databases = tmp_path / "databases"
    control_lists = databases / "control_lists"
    control_lists.mkdir(parents=True)
    (control_lists / "region_definitions.json").write_text(
        json.dumps(
            [
                {
                    "name": "Australia Group",
                    "acronym": "AG",
                    "regions": ["AU", "CA", "US"],
                },
                {
                    "name": "United Kingdom",
                    "acronym": "UK",
                    "regions": ["GB"],
                },
            ]
        ),
        encoding="utf-8",
    )
    direct_list = control_lists / "direct"
    direct_list.mkdir()
    (direct_list / "list_info.csv").write_text(
        "region_code\nBR\n",
        encoding="utf-8",
    )

    monkeypatch.setitem(server.CFG, "password_hash", None)
    monkeypatch.setitem(server.CFG, "require_local_auth", False)
    monkeypatch.setitem(server.CFG, "runs_dir", runs_dir)
    monkeypatch.setitem(server.CFG, "default_databases", str(databases))
    monkeypatch.setitem(
        server.CFG,
        "presets",
        [{"id": "test", "label": "Test", "config": {}}],
    )
    monkeypatch.setattr(server, "_missing_databases", lambda _preset: [])
    monkeypatch.setattr(server.threading.Thread, "start", lambda _thread: None)
    server.JOBS.clear()
    server.app.config.update(TESTING=True)
    yield server.app.test_client()
    server.JOBS.clear()


def _submit(client, **overrides):
    data = {
        "preset": "test",
        "label": "input-test",
        "sequence_text": ">short\nACGT\n",
    }
    data.update(overrides)
    return client.post(
        "/screen",
        data=data,
        environ_overrides={"REMOTE_ADDR": "127.0.0.1"},
    )


def _xlsx_file():
    book = openpyxl.Workbook()
    samples = book.active
    samples.title = "Samples"
    samples.append(["sample_id", "nucleotide", "note"])
    samples.append(["alpha", "ACGTACGTACGT", "first"])
    samples.append(["beta", "TTTTCCCCAAAA", "second"])
    other = book.create_sheet("Other")
    other.append(["name", "sequence"])
    other.append(["gamma", "GGGGAAAACCCC"])
    data = BytesIO()
    book.save(data)
    data.seek(0)
    return data


def _xls_file():
    book = xlwt.Workbook()
    sheet = book.add_sheet("Legacy")
    sheet.write(0, 0, "name")
    sheet.write(0, 1, "sequence")
    sheet.write(1, 0, "legacy")
    sheet.write(1, 1, "ACACACACACAC")
    data = BytesIO()
    book.save(data)
    data.seek(0)
    return data


def _long_xlsx_file():
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Long samples"
    sheet.append(["sample_id", "nucleotide"])
    for i in range(100):
        sheet.append([f"sample_{i + 1:03d}", "ACGT" * 25])
    data = BytesIO()
    book.save(data)
    data.seek(0)
    return data


def _long_header_xlsx_file():
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Long header"
    sheet.append([
        "Sample identifier header with spaces and over forty one characters",
        "Nucleotide sequence header with spaces and over forty one characters",
    ])
    sheet.append(["sample_001", "ACGT" * 25])
    data = BytesIO()
    book.save(data)
    data.seek(0)
    return data


def test_short_sequences_are_passed_to_commec(tmp_path, client):
    """The GUI leaves sequence-length handling to the screening pipeline."""
    response = _submit(
        client,
        # Older clients may still submit the removed field. It must not
        # restore the GUI's former length filter.
        skip_short="1",
    )

    assert response.status_code == 200
    run_dir = next(path for path in (tmp_path / "runs").iterdir() if path.is_dir())
    assert (run_dir / "input.fasta").read_text(encoding="utf-8") == ">short\nACGT\n"
    config = yaml.safe_load((run_dir / "config.used.yaml").read_text(encoding="utf-8"))
    assert config["databases"]["control_lists"]["regions"] == "all"

    status = client.get("/status").get_json()
    assert status["active"]["query_count"] == 1


def test_selected_regions_are_recorded_in_run_config(tmp_path, client):
    response = _submit(client, regions="US,CA")

    assert response.status_code == 200
    run_dir = next(path for path in (tmp_path / "runs").iterdir() if path.is_dir())
    config = yaml.safe_load((run_dir / "config.used.yaml").read_text(encoding="utf-8"))
    assert config["databases"]["control_lists"]["regions"] == "US,CA"


def test_hidden_single_country_alias_is_accepted(tmp_path, client):
    response = _submit(client, regions="UK")

    assert response.status_code == 200
    run_dir = next(path for path in (tmp_path / "runs").iterdir() if path.is_dir())
    config = yaml.safe_load((run_dir / "config.used.yaml").read_text(encoding="utf-8"))
    assert config["databases"]["control_lists"]["regions"] == "UK"


def test_unsupported_country_is_rejected(tmp_path, client):
    response = _submit(client, regions="AQ")

    assert response.status_code == 400
    assert response.get_json() == {
        "error": "Unknown regulatory jurisdiction: AQ.",
        "field": "settings",
    }
    assert list((tmp_path / "runs").iterdir()) == []


def test_unknown_region_is_rejected(tmp_path, client):
    response = _submit(client, regions="NOT_A_REGION")

    assert response.status_code == 400
    assert response.get_json() == {
        "error": "Unknown regulatory jurisdiction: NOT_A_REGION.",
        "field": "settings",
    }
    assert list((tmp_path / "runs").iterdir()) == []


def test_screen_errors_identify_their_gui_destination(client):
    missing_label = _submit(client, label="")
    assert missing_label.status_code == 400
    assert missing_label.get_json()["field"] == "label"

    missing_sequence = _submit(client, sequence_text="")
    assert missing_sequence.status_code == 400
    assert missing_sequence.get_json()["field"] == "sequence"

    unknown_preset = _submit(client, preset="missing")
    assert unknown_preset.status_code == 400
    assert unknown_preset.get_json()["field"] == "settings"

    started = _submit(client, label="first")
    assert started.status_code == 200
    busy = _submit(client, label="second")
    assert busy.status_code == 409
    assert busy.get_json()["field"] == "run"


def test_oversized_upload_is_a_sequence_error(client, monkeypatch):
    monkeypatch.setitem(server.app.config, "MAX_CONTENT_LENGTH", 1)

    response = _submit(client)

    assert response.status_code == 413
    assert response.get_json() == {
        "error": "Uploaded file is too large (max 200 MB).",
        "field": "sequence",
    }


def test_duplicate_run_label_is_a_label_error(tmp_path, client):
    retained = tmp_path / "runs" / "retained"
    retained.mkdir()
    (retained / "meta.json").write_text(
        json.dumps({"label": "duplicate"}), encoding="utf-8"
    )

    duplicate = _submit(client, label="duplicate")
    assert duplicate.status_code == 400
    assert duplicate.get_json() == {
        "error": "Can't accept run label: run label already chosen.",
        "field": "label",
    }


def test_config_exposes_grouped_region_choices(client):
    response = client.get(
        "/config",
        environ_overrides={"REMOTE_ADDR": "127.0.0.1"},
    )

    assert response.status_code == 200
    regions = response.get_json()["regions"]
    assert {
        "code": "US",
        "name": "United States",
        "group": False,
        "memberships": ["Australia Group"],
    } in regions
    assert {
        "code": "BR",
        "name": "Brazil",
        "group": False,
        "memberships": [],
    } in regions
    assert [region for region in regions if region["code"] == "AG"] == [
        {"code": "AG", "name": "Australia Group", "group": True},
    ]
    assert not any(region["code"] == "UK" for region in regions)
    assert {
        "code": "GB",
        "name": "United Kingdom",
        "group": False,
        "memberships": [],
    } in regions
    assert {
        region["code"] for region in regions if not region["group"]
    } == {"AU", "BR", "CA", "GB", "US"}


def test_gui_has_card_specific_message_destinations(client):
    response = client.get("/")

    assert response.status_code == 200
    html = response.get_data(as_text=True)
    for element_id in (
        "run-label-error",
        "sequence-error",
        "settings-error",
        "errmsg",
        "results-message",
    ):
        assert f'id="{element_id}"' in html
    assert "function showCardMessage(area, message, options = {})" in html
    assert 'const area = messageAreas[body.field] ? body.field : "run";' in html
    assert 'showCardMessage("results"' in html
    assert 'showCardMessage("run"' in html
    assert "function scrollToCard(card)" in html
    assert "const offset = topbar.getBoundingClientRect().height + 12;" in html


def test_gui_places_large_query_status_below_logo(client):
    response = client.get("/")

    assert response.status_code == 200
    html = response.get_data(as_text=True)
    brand_start = html.index('<div class="brand">')
    brand = html[brand_start:html.index('</div>', brand_start)]
    assert 'id="queue"' in brand
    assert 'class="queue-text"' in brand
    assert "margin: 0 0 0 8px" in html
    assert "width: 12px; height: 12px" in html
    assert "font: 700 var(--font-base)/1.2 var(--ui)" in html
    assert "s.active.query_count" in html
    assert '`Screening: ${label}${queries}`' in html


def test_gui_serves_bundled_report_font(client):
    response = client.get("/fonts/CrimsonPro.woff2")

    assert response.status_code == 200
    assert response.mimetype == "font/woff2"
    assert response.data[:4] == b"wOF2"


def test_xlsx_preview_lists_sheets_and_rows(client):
    response = client.post(
        "/spreadsheet-preview",
        data={"spreadsheet_file": (_xlsx_file(), "samples.xlsx")},
        content_type="multipart/form-data",
    )

    assert response.status_code == 200
    body = response.get_json()
    assert body["sheets"] == ["Samples", "Other"]
    assert body["sheet"] == "Samples"
    assert body["rows"][1][:2] == ["alpha", "ACGTACGTACGT"]


@pytest.mark.parametrize("filename,contents,sheet", [
    ("samples.xlsx", _xlsx_file, "Samples"),
    ("legacy.xls", _xls_file, "Legacy"),
    ("samples.csv", lambda: BytesIO(b"name,sequence\ncsv,CCCCAAAATTTT\n"), "Data"),
    ("samples.tsv", lambda: BytesIO(b"name\tsequence\ntsv\tAAAACCCCGGGG\n"), "Data"),
])
def test_tabular_upload_maps_selected_columns(tmp_path, client, filename, contents, sheet):
    response = _submit(
        client,
        sequence_text="",
        spreadsheet_file=(contents(), filename),
        spreadsheet_sheet=sheet,
        spreadsheet_nucleotide_column="1",
        spreadsheet_name_column="0",
        spreadsheet_skip_first_row="1",
    )

    assert response.status_code == 200
    run_dir = next(path for path in (tmp_path / "runs").iterdir() if path.is_dir())
    fasta = (run_dir / "input.fasta").read_text(encoding="utf-8")
    assert ">" in fasta
    assert "sequence" not in fasta


def test_tabular_upload_can_generate_names(client, tmp_path):
    response = _submit(
        client,
        sequence_text="",
        spreadsheet_file=(_xlsx_file(), "samples.xlsx"),
        spreadsheet_sheet="Other",
        spreadsheet_nucleotide_column="1",
        spreadsheet_name_column="",
        spreadsheet_skip_first_row="1",
    )

    assert response.status_code == 200
    run_dir = next(path for path in (tmp_path / "runs").iterdir() if path.is_dir())
    assert ">Other_row_2" in (run_dir / "input.fasta").read_text(encoding="utf-8")


def test_long_spreadsheet_preview_is_capped_but_all_rows_are_imported(client, tmp_path):
    preview = client.post(
        "/spreadsheet-preview",
        data={"spreadsheet_file": (_long_xlsx_file(), "long-samples.xlsx")},
        content_type="multipart/form-data",
    )

    assert preview.status_code == 200
    preview_data = preview.get_json()
    assert len(preview_data["rows"]) == server._TABULAR_PREVIEW_ROWS == 3
    assert preview_data["has_more_rows"] is True

    response = _submit(
        client,
        sequence_text="",
        spreadsheet_file=(_long_xlsx_file(), "long-samples.xlsx"),
        spreadsheet_sheet="Long samples",
        spreadsheet_nucleotide_column="1",
        spreadsheet_name_column="0",
        spreadsheet_skip_first_row="1",
    )

    assert response.status_code == 200
    run_dir = next(path for path in (tmp_path / "runs").iterdir() if path.is_dir())
    fasta = (run_dir / "input.fasta").read_text(encoding="utf-8")
    assert fasta.count(">sample_") == 100


def test_spreadsheet_header_is_passed_when_skip_first_row_is_not_selected(client, tmp_path):
    header_name = "Sample identifier header with spaces and over forty one characters"
    header_sequence = "Nucleotide sequence header with spaces and over forty one characters"
    assert len(header_sequence) > 41

    response = _submit(
        client,
        sequence_text="",
        spreadsheet_file=(_long_header_xlsx_file(), "long-header.xlsx"),
        spreadsheet_sheet="Long header",
        spreadsheet_nucleotide_column="1",
        spreadsheet_name_column="0",
    )

    assert response.status_code == 200
    run_dir = next(path for path in (tmp_path / "runs").iterdir() if path.is_dir())
    fasta = (run_dir / "input.fasta").read_text(encoding="utf-8")
    assert f">{header_name.replace(' ', '_')}\n{header_sequence}\n" in fasta
    assert ">sample_001\n" in fasta
